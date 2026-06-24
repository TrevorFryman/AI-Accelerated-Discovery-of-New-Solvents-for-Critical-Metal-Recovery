# -*- coding: utf-8 -*-
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SRC = r'C:\dev\DESs Selection.xlsx'
DST = r'C:\dev\ChemicalNames.xlsx'

with open(r'C:\dev\des_analysis\component_lookup.json', encoding='utf-8') as f:
    lookup = json.load(f)
with open(r'C:\dev\des_analysis\metalbinding_lookup.json', encoding='utf-8') as f:
    mb_lookup = json.load(f)
mb_lookup = {int(k): v for k, v in mb_lookup.items()}

src_wb = openpyxl.load_workbook(SRC, data_only=True)
dst_wb = openpyxl.Workbook()
dst_wb.remove(dst_wb.active)

HEADER_FILL = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=12)

def style_header(ws, row=1, ncols=None):
    ncols = ncols or ws.max_column
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def comp_info(name):
    e = lookup.get(name)
    if e is None:
        return ('', '', name)
    return (e['smiles'] or '', e['common_name'] or '', e['iupac_name'] or name)

# ---------- Pair-list sheets: MP Opt-RT, MP Opt_RDK-RT, MP Opt-WT, MP Opt_RDK-WT, Greenness ----------
PAIR_SHEETS = ['MP Opt-RT', 'MP Opt_RDK-RT', 'MP Opt-WT', 'MP Opt_RDK-WT']

for sheet_name in PAIR_SHEETS:
    src_ws = src_wb[sheet_name]
    headers = [c.value for c in src_ws[1]]
    dst_ws = dst_wb.create_sheet(sheet_name)

    # New header: Rank, Systematic Name #1, SMILES #1, Common Name #1, then same for #2, then remaining original columns (skip original c1/c2 cols 0,1)
    new_headers = ['Rank in List', 'Systematic Name #1', 'SMILES #1', 'Common Name #1',
                   'Systematic Name #2', 'SMILES #2', 'Common Name #2'] + headers[2:]
    dst_ws.append(new_headers)

    for i, row in enumerate(src_ws.iter_rows(min_row=2, max_row=26, values_only=True), start=1):
        c1, c2 = row[0], row[1]
        smi1, common1, iupac1 = comp_info(c1)
        smi2, common2, iupac2 = comp_info(c2)
        new_row = [i, iupac1, smi1, common1, iupac2, smi2, common2] + list(row[2:])
        dst_ws.append(new_row)

    style_header(dst_ws)
    dst_ws.freeze_panes = 'A2'
    widths = [10, 38, 30, 26, 38, 30, 26] + [16] * (len(headers) - 2)
    autosize(dst_ws, widths)

# ---------- Greenness sheet (already has SMILES; add common names) ----------
src_ws = src_wb['Greenness']
headers = [c.value for c in src_ws[1]]
dst_ws = dst_wb.create_sheet('Greenness')
new_headers = ['Rank', 'Systematic Name #1', 'SMILES #1', 'Common Name #1',
               'Systematic Name #2', 'SMILES #2', 'Common Name #2'] + headers[5:]
dst_ws.append(new_headers)
for row in src_ws.iter_rows(min_row=2, max_row=26, values_only=True):
    rank, c1, c2, smiles1, smiles2 = row[0], row[1], row[2], row[3], row[4]
    _, common1, iupac1 = comp_info(c1)
    _, common2, iupac2 = comp_info(c2)
    new_row = [rank, iupac1, smiles1, common1, iupac2, smiles2, common2] + list(row[5:])
    dst_ws.append(new_row)
style_header(dst_ws)
dst_ws.freeze_panes = 'A2'
autosize(dst_ws, [8, 38, 30, 26, 38, 30, 26] + [16] * (len(headers) - 5))

# ---------- Metal-Ligand_Ni+Co sheet ----------
src_ws = src_wb['Metal-Ligand_Ni+Co']
headers = [c.value for c in src_ws[1]]
dst_ws = dst_wb.create_sheet('Metal-Ligand_Ni+Co')
# original headers: Rank, Ligand_SMILES, LogK_mean_both, LogK_mean_Co, LogK_std_Co, LogK_mean_Ni, LogK_std_Ni,
#                    Delta_Ni_minus_Co, donor_atom_count, low_confidence, Common_name(=systematic name)
new_headers = ['Rank', 'Ligand SMILES', 'Systematic Name', 'Common Name',
               'LogK_mean_both', 'LogK_mean_Co', 'LogK_std_Co', 'LogK_mean_Ni', 'LogK_std_Ni',
               'Delta_Ni_minus_Co', 'donor_atom_count', 'low_confidence']
dst_ws.append(new_headers)
for row in src_ws.iter_rows(min_row=2, max_row=26, values_only=True):
    rank = row[0]
    smiles = row[1]
    systematic_name = row[10]
    common = mb_lookup.get(rank, {}).get('common_name', systematic_name)
    new_row = [rank, smiles, systematic_name, common, row[2], row[3], row[4], row[5], row[6], row[7], row[8], row[9]]
    dst_ws.append(new_row)
style_header(dst_ws)
dst_ws.freeze_panes = 'A2'
autosize(dst_ws, [8, 34, 50, 28, 14, 14, 12, 14, 12, 14, 12, 12])

# ---------- Component Glossary sheet (every unique compound, one row each) ----------
dst_ws = dst_wb.create_sheet('Component Glossary', 0)
dst_ws.append(['Systematic / Source Name', 'IUPAC Name', 'SMILES', 'Common Name', 'Molecular Formula', 'Molecular Weight', 'PubChem CID'])
for src_name in sorted(lookup.keys()):
    e = lookup[src_name]
    dst_ws.append([src_name, e['iupac_name'], e['smiles'], e['common_name'], e['molecular_formula'], e['molecular_weight'], e['cid']])
style_header(dst_ws)
dst_ws.freeze_panes = 'A2'
autosize(dst_ws, [55, 45, 32, 32, 16, 16, 12])

dst_ws2 = dst_wb.create_sheet('Metal-Binding Ligand Glossary', 1)
dst_ws2.append(['Rank', 'Systematic Name', 'SMILES', 'Common Name', 'PubChem CID'])
for rank in sorted(mb_lookup.keys()):
    e = mb_lookup[rank]
    dst_ws2.append([rank, e['systematic_name'], e['smiles'], e['common_name'], e['cid']])
style_header(dst_ws2)
dst_ws2.freeze_panes = 'A2'
autosize(dst_ws2, [8, 50, 34, 32, 12])

dst_wb.save(DST)
print('Saved', DST)
print('Sheets:', dst_wb.sheetnames)
