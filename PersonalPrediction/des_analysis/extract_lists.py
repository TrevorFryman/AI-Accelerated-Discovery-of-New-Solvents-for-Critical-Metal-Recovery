import openpyxl, json

wb = openpyxl.load_workbook(r'C:\dev\DESs Selection.xlsx', data_only=True)

def get_rows(sheet, min_row, max_row, cols):
    ws = wb[sheet]
    return [tuple(r) for r in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=cols[0], max_col=cols[1], values_only=True)]

data = {}

# MP_RDK-RT: Component1, Component2, X1, X2, Tmelt, predicted, uncertainty, similarity
ws = wb['MP Opt_RDK-RT']
rows = []
for r in ws.iter_rows(min_row=2, max_row=26, values_only=True):
    rows.append({'c1': r[0], 'c2': r[1], 'x1': r[2], 'x2': r[3], 'pred_Tmelt_K': r[5], 'uncertainty_K': r[6], 'similarity': r[7]})
data['MP_RDK_RT'] = rows

ws = wb['MP Opt_RDK-WT']
rows = []
for r in ws.iter_rows(min_row=2, max_row=26, values_only=True):
    rows.append({'c1': r[0], 'c2': r[1], 'x1': r[2], 'x2': r[3], 'pred_Tmelt_K': r[5], 'uncertainty_K': r[6], 'similarity': r[7]})
data['MP_RDK_WT'] = rows

ws = wb['Greenness']
rows = []
for r in ws.iter_rows(min_row=2, max_row=26, values_only=True):
    rows.append({'rank': r[0], 'c1': r[1], 'c2': r[2], 'smiles1': r[3], 'smiles2': r[4], 'x1': r[5], 'x2': r[6],
                 'Tmelt_K': r[7], 'type': r[8], 'doi': r[9], 'g1': r[10], 'g2': r[11], 'g_des': r[12]})
data['Greenness'] = rows

ws = wb['Metal-Ligand_Ni+Co']
rows = []
for r in ws.iter_rows(min_row=2, max_row=26, values_only=True):
    rows.append({'rank': r[0], 'smiles': r[1], 'logK_mean_both': r[2], 'logK_Co': r[3], 'logK_Ni': r[5],
                 'delta_Ni_Co': r[7], 'donor_atoms': r[8], 'low_confidence': r[9], 'name': r[10]})
data['MetalBinding'] = rows

with open(r'C:\dev\des_analysis\raw_lists.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, indent=2, default=str)

# collect unique component names needing SMILES (MP lists only have names)
names = set()
for r in data['MP_RDK_RT'] + data['MP_RDK_WT']:
    names.add(r['c1']); names.add(r['c2'])
for r in data['Greenness']:
    names.add(r['c1']); names.add(r['c2'])

with open(r'C:\dev\des_analysis\unique_names.json', 'w', encoding='utf-8') as f:
    json.dump(sorted(names), f, indent=2)

print(len(names), "unique component names")
for n in sorted(names):
    print(n)
